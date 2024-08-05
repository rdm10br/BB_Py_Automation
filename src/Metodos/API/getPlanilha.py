import pandas as pd
import openpyxl
import pyarrow
# import os
# print(os.getcwd())

# Acessando o arquivo
arq_excel = r'Planilhas\SALAS.xlsx'

# Lendo o arquivo
col = "ID"
col_status = 'STATUS'
df_map = pd.read_excel(arq_excel, sheet_name='salas')
total_lines = len(df_map)

col_plan2 = "ID_ORIGIN"
col_plan2_copy = 'ID_DESTINY'
col_plan2_status = 'STATUS'
df_map_plan2 = pd.read_excel(arq_excel, sheet_name='salaCopia')
total_lines_plan2 = len(df_map_plan2)

col_plan3_curso = "CURSO"
col_plan3_GA = 'GRANDE ÁREA'
df_map_plan3 = pd.read_excel(arq_excel, sheet_name='atividades')
total_lines_plan3 = len(df_map_plan3)

def getCell(index: int):
    """
    Function to get cell content from the 'salas' plan sheet and collum 'ID'
    of the excel file.

    Args:
        index (int): index of the line that you want to get from the excel file.

    Returns:
        Any: this function returns Any content that is on the cell
    """
    # Ajustando o índice para começar do zero
    index -= 1
    try :
    # Verificando se o índice está dentro do intervalo válido
        if 0 <= index < total_lines:
            # Obtendo o valor da célula na linha e coluna especificadas
            cell_value = df_map.at[index, col]
            return str(cell_value)
        else:
            return total_lines
    except Exception as e:
            print("index does not exist")
            
def getCell_status(index: int):
    """
    Function to get cell content from the 'salas' plan sheet and collum 'STATUS'
    of the excel file.

    Args:
        index (int): index of the line that you want to get from the excel file.

    Returns:
        Any: this function returns Any content that is on the cell
    """
    # Ajustando o índice para começar do zero
    index -= 1
    # Verificando se o índice está dentro do intervalo válido
    if 0 <= index < total_lines:
        # Obtendo o valor da célula na linha e coluna especificadas
        cell_value = df_map.at[index, col_status]
        return str(cell_value)
    else:
        return str(cell_value)

def getCell_plan2(index: int):
    """
    Function to get cell content from the 'salaCopia' plan sheet and collum 'ID_ORIGIN'
    of the excel file.

    Args:
        index (int): index of the line that you want to get from the excel file.

    Returns:
        Any: this function returns Any content that is on the cell
    """
    # Ajustando o índice para começar do zero
    index -= 1
    try :
    # Verificando se o índice está dentro do intervalo válido
        if 0 <= index < total_lines_plan2:
            # Obtendo o valor da célula na linha e coluna especificadas
            cell_value2 = df_map_plan2.at[index, col_plan2]
            return str(cell_value2)
        else:
            return total_lines_plan2
    except Exception as e:
            print("index does not exist")
            
def getCell_plan2_status(index: int):
    """
    Function to get cell content from the 'salaCopia' plan sheet and collum 'STATUS'
    of the excel file.

    Args:
        index (int): index of the line that you want to get from the excel file.

    Returns:
        Any: this function returns Any content that is on the cell
    """
    # Ajustando o índice para começar do zero
    index -= 1
    # Verificando se o índice está dentro do intervalo válido
    if 0 <= index < total_lines_plan2:
        # Obtendo o valor da célula na linha e coluna especificadas
        cell_value2 = df_map_plan2.at[index, col_plan2_status]
        return str(cell_value2)
    else:
        return str(cell_value2)
            
def getCell_copy_plan2(index: int):
    """
    Function to get cell content from the 'salaCopia' plan sheet and collum 'ID_DESTINY'
    of the excel file.

    Args:
        index (int): index of the line that you want to get from the excel file.

    Returns:
        Any: this function returns Any content that is on the cell
    """
    # Ajustando o índice para começar do zero
    index -= 1
    try :
    # Verificando se o índice está dentro do intervalo válido
        if 0 <= index < total_lines_plan2:
            # Obtendo o valor da célula na linha e coluna especificadas
            cell_value2 = df_map_plan2.at[index, col_plan2_copy]
            return str(cell_value2)
        else:
            return total_lines_plan2
    except Exception as e:
            print("index does not exist")
            
# def getCell_curso(index):
#     # Ajustando o índice para começar do zero
#     index -= 1
#     try :
#     # Verificando se o índice está dentro do intervalo válido
#         if 0 <= index < total_lines_plan3:
#             # Obtendo o valor da célula na linha e coluna especificadas
#             cell_value = df_map_plan3.at[index, col_plan3_curso]
#             return str(cell_value)
#         else:
#             return total_lines
#     except Exception as e:
#             print("index does not exist")
            
# def filter_GA(GA):
    
#     cursos_filtrados = df_map_plan3.loc[df_map_plan3[col_plan3_GA] == GA, 'CURSO']
    
#     return str(cursos_filtrados)
    
    
            
def writeOnExcel_Plan2(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salaCopia']
    
    col_status_plan2 = 'C' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan2}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'B' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1_Result(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'C' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1_Forum(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'D' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1_Desemp(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'E' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1_SOFIA(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'F' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1_FALE(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'G' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1_DESAFIO(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'H' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1_U1(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'I' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1_U2(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'J' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1_U3(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'K' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1_U4(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'L' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1_M1(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'M' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)

def writeOnExcel_Plan1_M2(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'N' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1_M3(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'O' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1_M4(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'P' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1_V1(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'Q' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1_V2(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'R' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1_V3(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'S' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1_V4(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'T' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1_B1(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'U' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1_B2(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'V' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)

def writeOnExcel_Plan1_B3(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'W' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1_B4(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'X' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1_A1(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'Y' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1_A2(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'Z' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1_A3(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'AA' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1_A4(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'AB' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1_AOL1(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'AC' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1_AOL2(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'AD' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1_AOL3(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'AE' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)

def writeOnExcel_Plan1_AOL4(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'AF' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)

def writeOnExcel_Plan1_AVALICAO(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'AG' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1_WEB(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'AH' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1_SOLICITE(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'AI' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)
    
def writeOnExcel_Plan1_SER(index, return_status):
    # Load an existing Excel workbook
    workbook = openpyxl.load_workbook(arq_excel)

    # Select the active sheet
    sheet = workbook['salas']
    
    col_status_plan1 = 'AJ' #COLUNA DE STATUS deve ser atribuida pela letra da coluna

    # Write data to the Excel sheet
    sheet[f'{col_status_plan1}{index+1}'] = return_status

    # Save the changes to the existing file
    workbook.save(arq_excel)

# index = 1
# for index in range(total_lines_plan2) :
#     index +=1
#     cell = getCell_plan2(index)
#     cell_copy = getCell_copy_plan2(index)
#     print(f'o número presente na celula é :{cell} na linha: {index+1} da cóluna : {col_plan2} e na coluna : {col_plan2_copy} está o ID de copia {cell_copy}')

# print(filter_GA(GA='COMUNICAÇÃO'))